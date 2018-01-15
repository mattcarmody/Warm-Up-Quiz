#!/usr/bin/env python3
# warmUp.py - Start the day with a mini quiz that adjusts to your
# understanding of the material over time.

import datetime
import numpy as np
import openpyxl
import random
#import sqlite3
#import tkinter as tk

# TODO: Web scrape from documentation pages to fill library
# TODO: Move to SQLite
# TODO: Add GUI

def callQuestion(jewelRow):
	# Identify first empty row in the calls log
	newCallsRow = sheet2.max_row + 1

	# Read in and print command (col D)
	print("\nWhat can you tell me about this prompt?")
	print(sheet1['D' + str(jewelRow)].value + '\n')
	input("Press enter when you're ready...")

	# Reveal the explanation (col E & F & G)
	print("")
	print(sheet1['E' + str(jewelRow)].value)
	print(sheet1['F' + str(jewelRow)].value)
	print(sheet1['G' + str(jewelRow)].value)
	
	# Prompt for and store user understanding on scale of l/m/h in main sheet (col B) & calls log (col C)
	understanding = input("\nHow well do you understand this? (l/m/h)")
	sheet1['B' + str(jewelRow)].value = understanding
	sheet2['C' + str(newCallsRow)].value = understanding
	
	# Update last called column (col A) & calls log (col B) with datetime
	sheet1['A' + str(jewelRow)].value = datetime.datetime.today()
	sheet2['B' + str(newCallsRow)].value = datetime.datetime.today()
	
	# Store command in calls log (col A) from main sheet (col B)
	sheet2['A' + str(newCallsRow)].value = sheet1['D' + str(jewelRow)].value

# Build probability model based on date last seen and understanding
def probModel():
	components = []
	
	# Loop through rows, weighting each question
	for i in range(2, maxrow + 1):
		# Increase probability of topics with a weaker understanding
		# If question has no history, make it highly probable
		understanding = sheet1['B' + str(i)].value
		if understanding == 'l':
			weight = 3
		elif understanding == 'm':
			weight = 2
		elif understanding == 'h':
			weight = 1
		else:
			weight = 1000000
		# Increase probability for topics that haven't been seen lately
		# If question has no history, make it highly probable
		lastSeen = sheet1['A' + str(i)].value
		try:
			diff = datetime.datetime.today().timestamp()-lastSeen.timestamp()
		except:	
			diff = 1000000
		# Combine weighting for understanding & datetime, then store it
		component = diff * weight
		components.append(component)
	# Use weights to create a probability np array
	prob = np.array(components, dtype=float) / sum(components)
	return prob

# Add GUI
# Choice of language/tool via radio buttons
# Enter integer for rounds (or radio buttons with an ALL option)
# FRESH: Show prompt and continue button (can later update to an entry form which will be documented for later analysis
# FRESH: Grid of values from spreadsheet and radio buttons for understanding level and continue button
# Loop as many times as needed
# You've finished screen with stats and a quit button 

# SQLite connection
#conn = sqlite3.connect('tutorial.db')
#c = conn.cursor()
# SELECT * FROM tool;  can a variable be used here?

# Prompt user for topic and length of quiz
tool = input("Which tool? Python or Excel\n")
quizCount = int(input("How many questions?\n"))

# Read in workbook and number of rows
wb = openpyxl.load_workbook("warmUp" + tool + ".xlsx")
sheet1 = wb.get_sheet_by_name("main")
sheet2 = wb.get_sheet_by_name("calls")
maxrow = sheet1.max_row
print("Data file currently has " + str(maxrow - 1) + " rows of data.")

# Call function to create probability model
prob = probModel()
print(prob)
rows = list(range(2, maxrow + 1))
print(rows)
jewelRows = np.random.choice(rows, size=quizCount, replace=False, p=prob)
print(jewelRows)

# Call question as many times as user requested
for i in range(quizCount):
    callQuestion(jewelRows[i])

# Save to spreadsheet file once at the end
wb.save("warmUp" + tool + ".xlsx")
